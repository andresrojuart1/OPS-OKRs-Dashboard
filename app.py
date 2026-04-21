"""
Ontop OPS OKRs Dashboard
Entry point: streamlit run app.py
"""

from datetime import datetime
import io
import os
import time

import openai
import openpyxl
import streamlit as st
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd

load_dotenv(override=True)

DEV_MODE = os.getenv("DEV_MODE", "false").lower() == "true"

st.set_page_config(
    page_title="Ontop OKRs — Operations",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---------------------------------------------------------------------------
# Theme CSS
# ---------------------------------------------------------------------------

st.markdown("""
<style>
:root {
    /* Brand */
    --ontop-purple: #7C5EFF;
    --ontop-coral:  #FF6B7A;

    /* Backgrounds with depth */
    --bg-primary:   #0A0E27;
    --bg-card:      #111633;
    --bg-input:     #1A2247;
    --bg-hover:     #1F2954;

    /* Text hierarchy */
    --text-primary:    #FFFFFF;
    --text-secondary:  #D1D5E0;
    --text-tertiary:   #8B94B3;
    --text-muted:      #5F6B8F;

    /* Borders */
    --border-color: #2A3555;

    /* Status system */
    --status-green:    #10B981;
    --status-yellow:   #FBBF24;
    --status-red:      #EF4444;
    --status-purple:   #A78BFA;

    /* Accent neutral */
    --accent-subtle: rgba(167, 139, 250, 0.08);
}

.stApp {
    background:
        radial-gradient(circle at top left,  rgba(124,94,255,0.25), transparent 32%),
        radial-gradient(circle at top right, rgba(255,107,122,0.15), transparent 28%),
        linear-gradient(180deg, #0A0E27 0%, #050814 100%);
    color: var(--text-primary);
}

[data-testid="stHeader"] { background: rgba(0,0,0,0); }

[data-testid="stSidebar"] {
    background:
        linear-gradient(180deg, rgba(124,94,255,0.12), rgba(10,14,39,0.96) 26%),
        #111633;
    border-right: 1px solid var(--border-color);
}

[data-testid="stSidebar"] * { color: var(--text-primary); }
[data-testid="stSidebarNav"]          { display: none; }
[data-testid="stSidebarNavSeparator"] { display: none; }

.block-container {
    padding-top: 1.75rem;
    padding-bottom: 3rem;
    max-width: 1100px;
}

h1,h2,h3 { color: var(--text-primary); letter-spacing: -0.02em; }
p, li, label, .stMarkdown, .stCaption { color: var(--text-secondary); }

/* Inputs */
.stTextInput input,
.stTextArea textarea,
.stSelectbox [data-baseweb="select"] > div,
.stNumberInput input {
    background: var(--bg-input) !important;
    border: 1px solid var(--border-color) !important;
    color: var(--text-primary) !important;
    border-radius: 14px !important;
    transition: all 0.2s ease !important;
}

.stTextInput input:focus,
.stTextArea textarea:focus,
.stSelectbox [data-baseweb="select"] > div:focus,
.stNumberInput input:focus {
    background: var(--bg-hover) !important;
    border-color: var(--status-purple) !important;
    box-shadow: 0 0 0 2px var(--accent-subtle) !important;
}

/* Primary button: purple→coral gradient, pill */
.stButton > button {
    background: linear-gradient(135deg, var(--ontop-purple), var(--ontop-coral));
    color: var(--text-primary);
    border: 0;
    border-radius: 999px;
    font-weight: 600;
    box-shadow: 0 10px 24px rgba(38,28,148,0.28);
}
.stButton > button:hover { filter: brightness(1.07); }

/* Secondary button — glassmorphism with better visibility */
.stButton > button[kind="secondary"] {
    background: rgba(255, 255, 255, 0.10) !important;
    border: 1px solid rgba(255, 255, 255, 0.20) !important;
    backdrop-filter: blur(12px) !important;
    color: #FFFFFF !important;
    border-radius: 12px !important;
    box-shadow: 0 2px 8px rgba(0,0,0,0.2) !important;
    transition: all 0.2s ease !important;
    font-weight: 500 !important;
}
.stButton > button[kind="secondary"]:hover {
    background: rgba(255, 255, 255, 0.18) !important;
    border-color: rgba(255, 255, 255, 0.35) !important;
    box-shadow: 0 4px 12px rgba(0,0,0,0.3) !important;
}

/* Tertiary — pill style */
.stButton > button[kind="tertiary"] {
    background: transparent !important;
    border: 1px solid rgba(255,255,255,0.10) !important;
    box-shadow: none !important;
    border-radius: 20px !important;
    padding: 6px 16px !important;
    transition: all 0.2s ease !important;
}
.stButton > button[kind="tertiary"]:hover {
    background: var(--accent-subtle) !important;
    border-color: rgba(255,255,255,0.15) !important;
}

/* Download button — match secondary glassmorphism */
[data-testid="stDownloadButton"] button {
    background: rgba(255, 255, 255, 0.08) !important;
    border: 1px solid rgba(255, 255, 255, 0.15) !important;
    backdrop-filter: blur(10px) !important;
    color: #FFFFFF !important;
    border-radius: 12px !important;
    font-weight: 600 !important;
    box-shadow: none !important;
    transition: all 0.2s ease !important;
    width: 100% !important;
}
[data-testid="stDownloadButton"] button:hover {
    background: rgba(255, 255, 255, 0.15) !important;
    border-color: rgba(255, 255, 255, 0.30) !important;
}

/* Destructive actions (Detects the delete material icon) */
div[data-testid="stButton"] button:has(span[data-testid="stIconMaterial"]:contains('delete')),
div[data-testid="stButton"] button:has(span[data-testid="stIconMaterial"]:contains('close')) {
    background: rgba(239, 68, 68, 0.1) !important;
    border: 1px solid rgba(239, 68, 68, 0.2) !important;
    color: #ef4444 !important;
}
div[data-testid="stButton"] button:has(span[data-testid="stIconMaterial"]:contains('delete')):hover,
div[data-testid="stButton"] button:has(span[data-testid="stIconMaterial"]:contains('close')):hover {
    background: rgba(239, 68, 68, 0.2) !important;
    border-color: rgba(239, 68, 68, 0.4) !important;
    color: #f87171 !important;
}

/* Glass/Elevation layer for cards */
.kr-card, .objective-card, .status-card {
    background: linear-gradient(
        180deg,
        rgba(255,255,255,0.02),
        rgba(255,255,255,0)
    ), var(--bg-card);
    border: 1px solid rgba(255,255,255,0.04);
    box-shadow: 0 4px 12px rgba(0,0,0,0.3);
}

/* Metrics */
div[data-testid="stMetric"] {
    background: linear-gradient(
        180deg,
        rgba(255,255,255,0.02),
        rgba(255,255,255,0)
    ), var(--bg-card);
    border: 1px solid rgba(255,255,255,0.04);
    border-radius: 18px;
    padding: 1rem;
    box-shadow: 0 4px 12px rgba(0,0,0,0.3);
}
[data-testid="stMetricValue"] { color: var(--ontop-coral) !important; }
[data-testid="stMetricLabel"] { color: var(--text-muted) !important; }

/* Status badges — integrated color system */
.status-success {
    background: rgba(16, 185, 129, 0.1);
    border: 1px solid rgba(16, 185, 129, 0.2);
    color: var(--status-green);
}

.status-warning {
    background: rgba(251, 191, 36, 0.1);
    border: 1px solid rgba(251, 191, 36, 0.2);
    color: var(--status-yellow);
}

.status-error {
    background: rgba(239, 68, 68, 0.1);
    border: 1px solid rgba(239, 68, 68, 0.2);
    color: var(--status-red);
}

.status-info {
    background: rgba(167, 139, 250, 0.1);
    border: 1px solid rgba(167, 139, 250, 0.2);
    color: var(--status-purple);
}

/* Alerts */
.stAlert { border-radius: 16px; border: 1px solid var(--border-color); }

/* Slider */
[data-testid="stSlider"] div[data-baseweb="slider"] div[role="slider"] {
    background: var(--ontop-coral) !important;
}

hr { border-color: var(--border-color) !important; }

::-webkit-scrollbar { width: 5px; }
::-webkit-scrollbar-track { background: #000; }
::-webkit-scrollbar-thumb { background: #2A2A3E; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: var(--ontop-coral); }

/* Sidebar brand */
.ontop-sidebar-brand { padding: 0.1rem 0 0.75rem; }
.ontop-sidebar-brand h1 { margin: 0.15rem 0 0; font-size: 1.45rem; line-height: 1; }
.ontop-sidebar-brand p  { margin: 0.3rem 0 0; color: var(--text-muted); font-size: 0.8rem; }

.ontop-sidebar-badge {
    display: inline-flex; align-items: center; gap: .35rem;
    padding: .28rem .55rem; border-radius: 999px;
    background: rgba(255,255,255,0.06); border: 1px solid rgba(255,255,255,0.08);
    color: #FFFFFF; font-size: .72rem; font-weight: 700;
    letter-spacing: .08em; text-transform: uppercase;
}

.ontop-sidebar-section-label {
    margin: .9rem 0 .4rem; color: var(--text-muted);
    text-transform: uppercase; letter-spacing: .08em;
    font-size: .72rem; font-weight: 700;
}

.ontop-sidebar-user {
    padding: .85rem; border-radius: 18px;
    border: 1px solid rgba(255,255,255,0.08);
    background:
        linear-gradient(180deg, rgba(255,255,255,0.02), rgba(255,255,255,0)),
        var(--bg-card);
    box-shadow: 0 2px 8px rgba(0,0,0,0.3);
    margin-top: .75rem; margin-bottom: .5rem;
    display: grid; grid-template-columns: 2.5rem 1fr; gap: .75rem; align-items: center;
}
.ontop-sidebar-avatar {
    width:2.5rem; height:2.5rem; border-radius:999px;
    display:flex; align-items:center; justify-content:center;
    background: linear-gradient(135deg, var(--ontop-purple), var(--ontop-coral));
    color:#fff; font-size:.9rem; font-weight:800;
    box-shadow: 0 2px 8px rgba(124,94,255,0.3);
}
.ontop-sidebar-user strong { display:block; color:#fff; font-size:.95rem; margin-bottom:.15rem; }
.ontop-sidebar-user span   { color:var(--text-secondary); font-size:.82rem; word-break:break-word; }
.ontop-sidebar-user-label  {
    display:block; color:var(--text-muted); font-size:.7rem;
    text-transform:uppercase; letter-spacing:.08em; margin-bottom:.18rem;
}

/* Nav link buttons in sidebar */
[data-testid="stSidebar"] .stButton > button {
    background: rgba(255,255,255,0.02) !important;
    border: 1px solid transparent !important;
    border-radius: 16px !important;
    color: var(--text-secondary) !important;
    font-size: .95rem !important;
    font-weight: 500 !important;
    text-align: left !important;
    padding: .55rem .82rem !important;
    justify-content: flex-start !important;
    box-shadow: none !important;
    width: 100% !important;
    min-height: 2.4rem !important;
    transition: all .12s ease !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    border-color: rgba(124,115,247,0.35) !important;
    background: rgba(255,255,255,0.05) !important;
    color: #fff !important;
    filter: none !important;
}

/* --- ONTOP BRAND DESIGN SYSTEM --- */
:root {
    --bg-main: #0B0B0F;
    --bg-surface: #111118;
    --bg-card: #161622;
    --accent-purple: #7A50F7;
    --accent-purple-dim: rgba(122, 80, 247, 0.15);
    --accent-indigo: #281782;
    --accent-teal: #2DD4BF;
    --accent-red: #EF4444;
    --text-primary: #FFFFFF;
    --text-secondary: rgba(255, 255, 255, 0.7);
    --text-muted: rgba(255, 255, 255, 0.45);
    --border-color: rgba(255, 255, 255, 0.06);
    --card-shadow: 0 4px 24px rgba(0, 0, 0, 0.35);
}

/* Global Ontop Layout */
.stApp {
    background-color: var(--bg-main) !important;
    font-family: 'Inter', -apple-system, sans-serif;
    color: var(--text-primary);
}

[data-testid="stSidebar"] {
    background-color: #08080C !important;
    border-right: 1px solid var(--border-color) !important;
}

.main-container {
    max-width: 1200px;
    margin: 0 auto;
    padding: 3rem 1.5rem;
}

/* Typography */
h1, h2, h3 {
/* Clean vertical rhythm for metric blocks */
hr {
    border: none;
    border-top: 1px solid rgba(255,255,255,0.08);
    margin: 12px 0 16px;
}

/* KPI Cards Styling */
.fintech-kpi-card {
    background: var(--bg-surface) !important;
    border: 1px solid var(--border-color) !important;
    border-radius: 16px !important;
    padding: 24px !important;
    box-shadow: var(--card-shadow) !important;
    transition: all 0.2s ease !important;
}
.fintech-kpi-card:hover {
    transform: translateY(-4px);
    border-color: var(--accent-purple-dim) !important;
    box-shadow: 0 12px 40px rgba(0, 0, 0, 0.4) !important;
}

/* ═══════════════════════════════════════════════════
   MASTER OBJECTIVE CARD
   ═══════════════════════════════════════════════════ */
div[data-testid="stVerticalBlock"]:has(> div:first-child div.okr-card-trigger) {
    background: var(--bg-surface) !important;
    border: 1px solid rgba(122, 80, 247, 0.15) !important;
    border-radius: 20px !important;
    padding: 2rem 2.25rem 1.75rem !important;
    margin-bottom: 2rem !important;
    box-shadow: var(--card-shadow), 0 0 40px rgba(122, 80, 247, 0.04) !important;
    transition: all 0.25s ease !important;
}
div[data-testid="stVerticalBlock"]:has(> div:first-child div.okr-card-trigger):hover {
    border-color: rgba(122, 80, 247, 0.3) !important;
    box-shadow: 0 8px 32px rgba(0, 0, 0, 0.45), 0 0 50px rgba(122, 80, 247, 0.06) !important;
}

/* Inner container resets inside objective card */
div[data-testid="stVerticalBlock"]:has(> div:first-child div.okr-card-trigger) div[data-testid="stVerticalBlock"] {
    background: transparent !important;
    border: none !important;
    box-shadow: none !important;
    padding: 0 !important;
    margin: 0 !important;
}

/* ═══════════════════════════════════════════════════
   KR BLOCKS (inside objective cards)
   Higher specificity than the inner reset above
   ═══════════════════════════════════════════════════ */
div[data-testid="stVerticalBlock"]:has(> div:first-child div.okr-card-trigger)
div[data-testid="stVerticalBlock"]:has(> div:first-child div.kr-block-trigger) {
    background: rgba(255, 255, 255, 0.02) !important;
    border: 1px solid rgba(255, 255, 255, 0.06) !important;
    border-radius: 14px !important;
    padding: 18px 20px !important;
    margin-bottom: 10px !important;
    transition: all 0.2s ease !important;
}
div[data-testid="stVerticalBlock"]:has(> div:first-child div.okr-card-trigger)
div[data-testid="stVerticalBlock"]:has(> div:first-child div.kr-block-trigger):hover {
    background: rgba(255, 255, 255, 0.04) !important;
    border-color: rgba(255, 255, 255, 0.12) !important;
    box-shadow: 0 2px 12px rgba(0, 0, 0, 0.2) !important;
}

/* Inner container resets within KR blocks */
div[data-testid="stVerticalBlock"]:has(> div:first-child div.kr-block-trigger) div[data-testid="stVerticalBlock"] {
    background: transparent !important;
    border: none !important;
    box-shadow: none !important;
    padding: 0 !important;
    margin: 0 !important;
}

/* ═══════════════════════════════════════════════════
   BUTTONS
   ═══════════════════════════════════════════════════ */
div.stButton > button {
    border-radius: 10px !important;
    transition: all 0.2s ease !important;
}
div.stButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 12px rgba(0, 0, 0, 0.2) !important;
}
div.stButton > button[kind="secondary"] {
    background: rgba(255, 255, 255, 0.08) !important;
    border: 1px solid rgba(255, 255, 255, 0.1) !important;
    color: #FFFFFF !important;
    font-weight: 600 !important;
    font-size: 13px !important;
}
div.stButton > button[kind="secondary"]:hover {
    background: rgba(255, 255, 255, 0.15) !important;
    border-color: var(--accent-purple) !important;
}

/* Expander styling for history */
div[data-testid="stExpander"] {
    border: 1px solid rgba(255, 255, 255, 0.06) !important;
    border-radius: 10px !important;
    background: rgba(255, 255, 255, 0.015) !important;
}
div[data-testid="stExpander"] summary {
    color: rgba(255, 255, 255, 0.5) !important;
    font-size: 13px !important;
    font-weight: 600 !important;
}
</style>

""", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
from auth import require_login, get_user, logout
from sheets_cached import (
    load_objectives_cached as load_objectives,
    load_key_results_cached as load_key_results,
    load_updates_cached as load_updates,
    clear_sheets_cache,
    load_weekly_notes_cached,
    load_weekly_charts_cached,
)
from sheets import (
    seed_if_empty,
    compute_progress, create_objective,
    get_week_number, get_weekly_note, save_weekly_note,
    undo_last_import,
    get_weekly_charts, upload_charts_to_drive, delete_chart_from_drive,
    download_drive_file,
)
from components.sidebar import render_sidebar, SUB_TEAMS
from components.objective_card import render_objective_card
from pdf_parser import parse_okr_pdf_with_ai, render_pdf_preview_and_confirm
from html_export import generate_html_report
from observability import logger, track_action, handle_error, render_last_action, render_activity_log

# ---------------------------------------------------------------------------
for key, default in [
    ("updating_kr",       None),
    ("selected_team",     "All"),
    ("selected_quarter",  "Q2 2026"),
    ("ai_dialog_stale",   True),
    ("show_pdf_import",   False),
    ("last_sync_time",    None),
]:
    if key not in st.session_state:
        st.session_state[key] = default

# Initialize presentation mode for each team
TEAMS = ["All", "New Initiatives", "FinOps", "AI Experience", "AI Monetization", "Security & Compliance Ops"]
for team in TEAMS:
    key = f"presentation_mode_{team}"
    if key not in st.session_state:
        st.session_state[key] = False

QUARTERS = ["Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026"]

# ---------------------------------------------------------------------------
# Excel template
# ---------------------------------------------------------------------------

def _generate_template_excel(quarter: str, krs_info: list = None, objectives_df: pd.DataFrame = None,
                               updates_df: pd.DataFrame = None) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Status {quarter}"

    headers = [
        "Objective",
        "Key Result",
        "Target",
        "Current Value",
        "Confidence (1–5)",
        "What happened this week?",
        "Blockers / Dependencies",
    ]

    hdr_fill = PatternFill(start_color="1E2140", end_color="1E2140", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True)

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    rows = []
    if krs_info:
        for k in krs_info:
            # Handle both dict structures: direct dict and nested dict with "kr" key
            kr_data = k.get("kr") if "kr" in k else k

            # Get values from the appropriate source
            val = k.get("val") if "val" in k else kr_data.get("current_value", 0)
            tgt = kr_data.get("target", 0)
            title = k.get("title") if "title" in k else kr_data.get("title", "")
            unit = k.get("unit") if "unit" in k else kr_data.get("unit", "")
            kr_id = str(kr_data.get("id", ""))

            # Get Objective from objectives_df
            obj_title = ""
            if not objectives_df.empty and "id" in objectives_df.columns:
                obj_id = str(kr_data.get("objective_id", ""))
                obj_rows = objectives_df[objectives_df["id"].astype(str) == obj_id]
                if not obj_rows.empty:
                    obj_title = str(obj_rows.iloc[0].get("title", ""))

            # Get Confidence, Blockers, and Week Notes from updates_df (most recent update for this KR)
            confidence = ""
            blockers = ""
            week_notes = ""
            if not updates_df.empty and kr_id:
                kr_updates = updates_df[updates_df["kr_id"].astype(str) == kr_id]
                if not kr_updates.empty:
                    # Get the most recent update
                    latest_update = kr_updates.sort_values("updated_at", ascending=False).iloc[0]
                    confidence = str(latest_update.get("confidence", "")) if latest_update.get("confidence") else ""
                    blockers = str(latest_update.get("blockers", ""))[:100] if latest_update.get("blockers") else ""
                    week_notes = str(latest_update.get("week_notes", ""))[:100] if latest_update.get("week_notes") else ""

            # Formatting logic for Excel
            fmt = str(kr_data.get("value_format", "number")).lower() if "value_format" in kr_data else "number"

            if fmt == "percentage":
                val_str = f"{val}%"
                tgt_str = f"{tgt}%"
            elif fmt == "currency":
                val_str = f"${val:,.2f}"
                tgt_str = f"${tgt:,.2f}"
            else:
                val_str = str(val)
                tgt_str = f"{tgt} {unit}".strip() if unit else str(tgt)

            rows.append([
                obj_title,
                title,
                tgt_str,
                val_str,
                confidence if confidence else "",
                week_notes,
                blockers if blockers else ""
            ])
    else:
        # Fallback to empty rows / instructions
        rows.append(["[No data found for this selection]", "", "", "", "", "", ""])

    fill_a = PatternFill(start_color="13152A", end_color="13152A", fill_type="solid")
    fill_b = PatternFill(start_color="0D0E1A", end_color="0D0E1A", fill_type="solid")
    body_font = Font(color="FFFFFF")

    for i, row in enumerate(rows):
        ws.append(row)
        fill = fill_a if i % 2 == 0 else fill_b
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.font = body_font

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        col_letter = col[0].column_letter
        # Objective column (A) needs more width for full text
        max_width = 80 if col_letter == "A" else 60
        ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, max_width))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# AI Update dialog (OpenAI)
# ---------------------------------------------------------------------------

@st.dialog("AI Weekly Update")
def _ai_update_dialog() -> None:
    quarter   = st.session_state.get("selected_quarter", "Q2 2026")
    sub_team  = st.session_state.get("selected_team",    "All")
    krs_info  = st.session_state.get("_krs_for_ai",      [])
    cache_key = f"_ai_text_{quarter}_{sub_team}"

    if st.session_state.get("ai_dialog_stale", True) or cache_key not in st.session_state:
        team_label = sub_team if sub_team != "All" else "Operations (all teams)"

        lines = []
        for kr in krs_info:
            lines.append(
                f"- {kr['title']}: {kr['current_value']}/{kr['target']} {kr['unit']} "
                f"({kr['pct']:.0f}%)"
            )
            if kr.get("last_notes"):
                lines.append(f"  Last week: {kr['last_notes']}")
            if kr.get("last_blockers"):
                lines.append(f"  Blockers: {kr['last_blockers']}")
            if kr.get("last_confidence"):
                lines.append(f"  Confidence: {kr['last_confidence']}/5")
        krs_text = "\n".join(lines) or "No KRs available."

        prompt = (
            f"You are an operations assistant at Ontop. Based on the current OKRs state "
            f"for the team {team_label} for {quarter}, generate an executive summary in English of "
            f"maximum 300 words that includes: 1) Overall quarter status, "
            f"2) Key achievements, 3) Critical risks and blockers, "
            f"4) Concrete recommendations for next week.\n\n"
            f"Current KR data:\n{krs_text}"
        )

        with st.spinner("Generating AI summary..."):
            try:
                client = openai.OpenAI(api_key=st.secrets["OPENAI_API_KEY"])
                response = client.chat.completions.create(
                    model="gpt-4o",
                    messages=[{"role": "user", "content": prompt}],
                    max_tokens=800,
                )
                st.session_state[cache_key] = response.choices[0].message.content
                st.session_state["ai_dialog_stale"] = False
            except Exception as exc:
                st.error(f"Error calling OpenAI API: {exc}")
                return

    text = st.session_state.get(cache_key, "")
    st.markdown(text)
    st.divider()
    st.caption("Copy full text:")
    st.code(text, language=None)


# ---------------------------------------------------------------------------
# Add Objective dialog
# ---------------------------------------------------------------------------

@st.dialog("New Objective")
def add_objective_dialog(sub_team: str, quarter: str) -> None:
    _subteams = [t for t in SUB_TEAMS if t != "All"]

    if sub_team == "All":
        actual_team = st.selectbox("Sub-team", _subteams)
    else:
        actual_team = sub_team

    title = st.text_area(
        "Objective title",
        placeholder="Ex: Scale Financial Products into Reliable ARR Contributors…",
    )
    st.caption(f"Sub-team: **{actual_team}** · Quarter: **{quarter}**")

    c1, c2 = st.columns(2)
    with c1:
        if st.button("Cancel", width="stretch", key="add_obj_cancel"):
            st.rerun()
    with c2:
        if st.button("Create", width="stretch", type="primary"):
            if title.strip():
                create_objective(title.strip(), actual_team, quarter)
                st.rerun()
            else:
                st.error("Title cannot be empty.")


# ---------------------------------------------------------------------------
# Login page
# ---------------------------------------------------------------------------

def render_login_page() -> None:
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("""
        <div style="text-align:center;padding:80px 0 48px;">
            <div style="margin-bottom:24px;">
                <span style="font-size:38px;">🔺</span>
                <span style="font-size:38px;font-weight:900;color:#fff;letter-spacing:-1px;"> ontop</span>
            </div>
            <div style="font-size:22px;font-weight:700;color:#fff;margin-bottom:6px;">
                Operations OKRs
            </div>
            <div style="font-size:14px;color:#6B6B7E;margin-bottom:44px;">
                Sign in with your @getontop.com account
            </div>
        </div>
        """, unsafe_allow_html=True)

        if DEV_MODE:
            st.warning("DEV MODE — OAuth bypassed")
            if st.button("Enter as mrojas@getontop.com", width="stretch"):
                st.session_state["dev_authenticated"] = True
                st.session_state["user"] = {
                    "email": "mrojas@getontop.com", "name": "Andrés Rojas",
                    "given_name": "Andrés", "picture": "",
                }
                st.rerun()
            return

        if st.button("Sign in with Google", width="stretch"):
            st.login("google")


# ---------------------------------------------------------------------------
# Header
# ---------------------------------------------------------------------------

def render_header(objectives_df, krs_df, updates_df, selected_team, krs_info, krs_info_all=None) -> None:
    """Dashboard header with system-level actions and KPI summaries."""
    selected_quarter = st.session_state.get("selected_quarter", "Q2 2026")

    # Use krs_info_all for Excel export (complete data), fallback to krs_info if not provided
    krs_info_for_export = list(krs_info_all.values()) if krs_info_all else krs_info

    col_l, col_r = st.columns([1.5, 3.5])
    with col_l:
        sub = f" · {selected_team}" if selected_team != "All" else ""
        st.markdown(f"""
        <div style="margin-bottom:8px;">
            <div style="font-size:24px;font-weight:800;color:#fff;line-height:1.1;">
                Operations OKRs{sub}
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # Quarter selector in a narrower container
        with st.container():
            c_q, _ = st.columns([1.5, 1])
            with c_q:
                st.selectbox(
                    "Quarter",
                    QUARTERS,
                    index=QUARTERS.index(selected_quarter) if selected_quarter in QUARTERS else 1,
                    key="selected_quarter",
                    label_visibility="collapsed",
                )

    with col_r:
        st.markdown("<div style='height:32px;'></div>", unsafe_allow_html=True)

        # Compact header buttons: Sync • Excel • Report • ⬆ PDF • [Present/Edit if subteam]
        selected_week = st.session_state.get("selected_week", get_week_number())
        show_pdf_import = selected_team != "All"

        # Calculate columns: Sync | Excel | Report | PDF | [Present if subteam]
        num_cols = 5 if show_pdf_import else 4
        cols = st.columns(num_cols)

        # 1. Sync
        with cols[0]:
            if st.button("Sync", icon=":material/refresh:", key="hdr_sync", width="stretch", type="secondary", help="Refresh data from Sheets"):
                clear_sheets_cache()
                st.session_state["last_sync_time"] = datetime.now()
                track_action("Synced data")
                st.toast("Data synchronized", icon="✅")
                st.rerun()

        # 2. Excel
        with cols[1]:
            excel_bytes = _generate_template_excel(selected_quarter, krs_info_for_export, objectives_df, updates_df)
            st.download_button(
                label="Excel",
                icon=":material/download:",
                data=excel_bytes,
                file_name=f"OKRs_{selected_quarter}_W{selected_week}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="hdr_excel",
                width="stretch",
                type="secondary",
                help="Download as Excel"
            )

        # 3. Report
        with cols[2]:
            notes_df = load_weekly_notes_cached()
            charts_df = load_weekly_charts_cached()
            html_content = generate_html_report(
                objectives_df=objectives_df,
                krs_df=krs_df,
                updates_df=updates_df,
                notes_df=notes_df,
                quarter=selected_quarter,
                charts_df=charts_df,
            )
            st.download_button(
                label="Report",
                icon=":material/file_download:",
                data=html_content.encode('utf-8'),
                file_name=f"OKRs_{selected_quarter}_W{selected_week}.html",
                mime="text/html",
                key="hdr_report",
                width="stretch",
                type="secondary",
                help="Download HTML report"
            )

        # 4. PDF (Upload)
        with cols[3]:
            if st.button("PDF", icon=":material/upload:", key="pdf_import_btn_hdr", width="stretch", type="secondary", help="Upload from PDF"):
                st.session_state["show_pdf_import"] = True

        # 5. Present/Edit Mode toggle (only for subteams)
        if show_pdf_import:
            with cols[4]:
                presentation_key = f"presentation_mode_{selected_team}"
                is_presentation = st.session_state.get(presentation_key, False)
                button_label = "Edit" if is_presentation else "Present"
                button_icon = ":material/edit:" if is_presentation else ":material/tv:"
                button_help = "Exit presentation mode" if is_presentation else "Enter presentation mode"

                if st.button(button_label, icon=button_icon, key="hdr_presentation", width="stretch", type="secondary", help=button_help):
                    st.session_state[presentation_key] = not st.session_state.get(presentation_key, False)
                    st.rerun()

    # Metrics from krs_info
    num_objs = len(objectives_df[objectives_df["quarter"] == selected_quarter]) if not objectives_df.empty else 0
    if selected_team != "All":
        num_objs = len(objectives_df[(objectives_df["quarter"] == selected_quarter) & (objectives_df["sub_team"] == selected_team)])

    # Dynamic Progress Threshold (Linear target per quarter week)
    selected_week = st.session_state.get("selected_week", 1)
    q_starts = {"Q1 2026": 1, "Q2 2026": 14, "Q3 2026": 27, "Q4 2026": 40}
    start_wk = q_starts.get(selected_quarter, 1)
    weeks_elapsed = max(1, selected_week - start_wk + 1)
    expected_pct = (weeks_elapsed / 13.0) * 100

    total_krs = len(krs_info)
    if krs_info:
        # ONLY count KRs that have at least one update record
        active_krs = [k for k in krs_info if k.get("has_updates", False)]
        at_risk_count = sum(1 for k in active_krs if k["pct"] < expected_pct)
        on_track_count = sum(1 for k in active_krs if k["pct"] >= expected_pct)
        avg_prog = sum(k["pct"] for k in krs_info) / total_krs
    else:
        at_risk_count = 0
        on_track_count = 0
        avg_prog = 0.0
        
    def _get_status_color(ratio, inverse=False):
        if inverse:
            # Low is good (for At Risk) — fewer at-risk = better
            if ratio <= 0.1: return "#10b981" # Green — 0-10% at risk
            if ratio <= 0.25: return "#fbbf24" # Yellow — 10-25% at risk
            if ratio <= 0.5: return "#f59e0b" # Orange — 25-50% at risk
            return "#ef4444" # Red — 50%+ at risk
        else:
            # High is good
            if ratio >= 0.8: return "#10b981" # Green
            if ratio >= 0.5: return "#fbbf24" # Yellow
            if ratio >= 0.25: return "#f59e0b" # Orange
            return "#ef4444" # Red

    prog_ratio = avg_prog / 100.0
    ot_ratio = on_track_count / total_krs if total_krs > 0 else 0
    ar_ratio = at_risk_count / total_krs if total_krs > 0 else 0

    prog_kpi_color = _get_status_color(prog_ratio)
    # On Track is always GREEN (if there are on-track KRs, that's good)
    ot_kpi_color = "#10b981" if on_track_count > 0 else "#6B6B7E"
    # At Risk scales: less at-risk = greener, more at-risk = redder
    ar_kpi_color = _get_status_color(ar_ratio, inverse=True)
    # KPI Cards
    st.markdown(f"""
    <div style="display:flex; gap:16px; margin-bottom:32px; flex-wrap:wrap;">
        <div class="fintech-kpi-card" style="flex:1; min-width: 140px; border-bottom: 2px solid var(--border-color);">
            <div style="color:var(--text-secondary); font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:0.08em; opacity:0.7;">Objectives</div>
            <div style="color:#FFFFFF; font-size:32px; font-weight:700; line-height:1; margin-top:12px; letter-spacing:-0.02em;">{num_objs}</div>
            <div style="color:var(--text-muted); font-size:12px; font-weight:500; margin-top:8px;">Total Active</div>
        </div>
        <div class="fintech-kpi-card" style="flex:1; min-width: 140px; border-bottom: 2px solid var(--border-color);">
            <div style="color:var(--text-secondary); font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:0.08em; opacity:0.7;">Key Results</div>
            <div style="color:#FFFFFF; font-size:32px; font-weight:700; line-height:1; margin-top:12px; letter-spacing:-0.02em;">{total_krs}</div>
            <div style="color:var(--text-muted); font-size:12px; font-weight:500; margin-top:8px;">Tracked Items</div>
        </div>
        <div class="fintech-kpi-card" style="flex:1; min-width: 140px; border-bottom: 2px solid {prog_kpi_color}66;">
            <div style="color:var(--text-secondary); font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:0.08em; opacity:0.7;">Execution Progress</div>
            <div style="color:{prog_kpi_color}; font-size:32px; font-weight:700; line-height:1; margin-top:12px; letter-spacing:-0.02em;">{avg_prog:.0f}%</div>
            <div style="color:var(--text-muted); font-size:12px; font-weight:500; margin-top:8px;">Overall Completion</div>
        </div>
        <div class="fintech-kpi-card" style="flex:1; min-width: 140px; border-bottom: 2px solid {ot_kpi_color}66;">
            <div style="color:var(--text-secondary); font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:0.08em; opacity:0.7;">KRs On Track</div>
            <div style="color:{ot_kpi_color}; font-size:32px; font-weight:700; line-height:1; margin-top:12px; letter-spacing:-0.02em;">{on_track_count}</div>
            <div style="color:var(--text-muted); font-size:12px; font-weight:500; margin-top:8px;">Healthy Items</div>
        </div>
        <div class="fintech-kpi-card" style="flex:1; min-width: 140px; border-bottom: 2px solid {ar_kpi_color}66;">
            <div style="color:var(--text-secondary); font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:0.08em; opacity:0.7;">KRs At Risk</div>
            <div style="color:{ar_kpi_color}; font-size:32px; font-weight:700; line-height:1; margin-top:12px; letter-spacing:-0.02em;">{at_risk_count}</div>
            <div style="color:var(--text-muted); font-size:12px; font-weight:500; margin-top:8px;">Needs Attention</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<hr style='margin:10px 0 16px;'>", unsafe_allow_html=True)

    # TODO: AI Weekly Summary - Disabled for now, activate when ready
    # Generate "What Changed This Week" with AI (hybrid: once per session + refresh button)
    # ai_summary_key = f"ai_summary_{selected_team}_{selected_quarter}_{selected_week}"
    #
    # # Check if we already have a summary cached in session
    # if ai_summary_key not in st.session_state:
    #     st.session_state[ai_summary_key] = None
    #
    # col_title, col_refresh = st.columns([0.95, 0.05])
    # with col_title:
    #     st.markdown('<div style="font-size:13px; font-weight:800; color:rgba(255,255,255,0.5); text-transform:uppercase; letter-spacing:0.1em; padding-bottom:8px; margin-bottom:12px;">📝 What Changed This Week</div>', unsafe_allow_html=True)
    #
    # with col_refresh:
    #     if st.button("🔄", key=f"refresh_ai_{selected_team}_{selected_week}", help="Refresh summary"):
    #         st.session_state[ai_summary_key] = None
    #         st.rerun()
    #
    # # Generate or use cached summary
    # if st.session_state[ai_summary_key] is None:
    #     with st.spinner("Generating summary..."):
    #         try:
    #             # Collect narratives + dependencies from current week AND last week (for trends)
    #             current_week_updates = updates_df[updates_df["week_number"] == selected_week] if not updates_df.empty else pd.DataFrame()
    #             last_week_updates = updates_df[updates_df["week_number"] == selected_week - 1] if not updates_df.empty else pd.DataFrame()
    #
    #             narratives = []
    #             dependencies = []
    #
    #             # Current week narratives
    #             if not current_week_updates.empty:
    #                 narratives = current_week_updates["week_notes"].dropna().tolist()
    #
    #             # Get dependencies from current week
    #             for _, kr_row in krs_df.iterrows():
    #                 kr_id = str(kr_row["id"])
    #                 kr_updates = current_week_updates[current_week_updates["kr_id"].astype(str) == kr_id]
    #                 if not kr_updates.empty:
    #                     deps = kr_updates["blockers"].dropna().unique().tolist()
    #                     if deps:
    #                         dependencies.extend(deps)
    #
    #             # Calculate week-over-week changes
    #             current_avg_pct = current_week_updates["new_value"].astype(float).mean() if not current_week_updates.empty else 0
    #             last_avg_pct = last_week_updates["new_value"].astype(float).mean() if not last_week_updates.empty else 0
    #             pct_change = current_avg_pct - last_avg_pct
    #
    #             # Build prompt for AI
    #             prompt = f"""Summarize what changed this week in the Operations team OKRs in a brief, executive-friendly paragraph (2-3 sentences max).
    #
    # Current Week: Week {selected_week}
    # Team: {selected_team}
    # Quarter: {selected_quarter}
    #
    # Trend: {'↑ Up' if pct_change > 0 else '↓ Down' if pct_change < 0 else '→ Stable'} {abs(pct_change):.1f}pp from last week
    #
    # Key Updates This Week:
    # {chr(10).join([f"- {n}" for n in narratives[:5]]) if narratives else "No updates recorded"}
    #
    # Dependencies/Blockers:
    # {chr(10).join([f"- {d}" for d in dependencies[:3]]) if dependencies else "None"}
    #
    # Generate a concise executive summary like:
    # "Aura CSAT dropped 1.1pp due to Quick spike (8→30). Troubleshooting being added. Card fraud up 4x, coaching sessions scheduled."
    #
    # If no updates, say: "No updates recorded for {selected_team} this week."""
    #
    #             client = openai.OpenAI()
    #             response = client.chat.completions.create(
    #                 model="gpt-4",
    #                 messages=[{"role": "user", "content": prompt}],
    #                 temperature=0.7,
    #                 max_tokens=150
    #             )
    #
    #             summary_text = response.choices[0].message.content.strip()
    #             st.session_state[ai_summary_key] = summary_text
    #         except Exception as e:
    #             st.session_state[ai_summary_key] = f"Could not generate summary: {str(e)}"
    #             track_action("AI summary generation failed", detail=str(e))
    #
    #     # Display the summary
    #     summary_text = st.session_state[ai_summary_key]
    #     if summary_text:
    #         st.markdown(f'<div style="background:rgba(255,255,255,0.03); padding:12px 16px; border-radius:8px; border-left:3px solid #7A50F7; color:rgba(255,255,255,0.8); font-size:14px; line-height:1.5;">{summary_text}</div>', unsafe_allow_html=True)
    #     else:
    #         st.markdown(f'<div style="color:rgba(255,255,255,0.4); font-size:13px; font-style:italic;">No updates recorded for {selected_team} this week.</div>', unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

def render_dashboard() -> None:
    if not DEV_MODE and "user" not in st.session_state:
        st.session_state["user"] = get_user()

    try:
        seed_if_empty()
    except Exception as exc:
        import traceback
        st.error(f"[{type(exc).__name__}] {exc}")
        st.code(traceback.format_exc())
        st.stop()

    objectives_df = load_objectives()
    if not objectives_df.empty and "sub_team" in objectives_df.columns:
        objectives_df["sub_team"] = objectives_df["sub_team"].replace({"Growth Initiatives": "New Initiatives"})
        
    krs_df        = load_key_results()
    updates_df    = load_updates()
    
    # --- MULTI-LEVEL DATA CONSISTENCY ---
    # 1. Filter Objectives
    if not objectives_df.empty:
        objectives_df = objectives_df[
            (objectives_df["id"].astype(str).replace("nan", "").str.strip().str.len() > 0) &
            (objectives_df["title"].astype(str).replace("nan", "").str.strip().str.len() > 0)
        ].copy()
        
    # 2. Filter KRs (Must have active Objective AND a title)
    if not krs_df.empty and not objectives_df.empty:
        valid_obj_ids = set(objectives_df["id"].astype(str).unique())
        krs_df = krs_df[
            (krs_df["id"].astype(str).replace("nan", "").str.strip().str.len() > 0) & 
            (krs_df["title"].astype(str).replace("nan", "").str.strip().str.len() > 0) &
            (krs_df["objective_id"].astype(str).isin(valid_obj_ids))
        ].copy()
        
    # 3. Filter Updates (Must have active KR)
    if not updates_df.empty and not krs_df.empty:
        valid_kr_ids = set(krs_df["id"].astype(str).unique())
        updates_df = updates_df[
            (updates_df["id"].astype(str).replace("nan", "").str.strip().str.len() > 0) & 
            (updates_df["kr_id"].astype(str).isin(valid_kr_ids))
        ].copy()



    
    # --- DYNAMIC WEEK INITIALIZATION ---
    if "selected_week" not in st.session_state:
        if not updates_df.empty and "week_number" in updates_df.columns:
            latest_data_wk = pd.to_numeric(updates_df["week_number"], errors="coerce").max()
            if pd.notna(latest_data_wk) and latest_data_wk > 0:
                st.session_state["selected_week"] = int(latest_data_wk)

            else:
                st.session_state["selected_week"] = get_week_number()
        else:
            st.session_state["selected_week"] = get_week_number()
    
    # --- PERSISTENT DIALOGS ---
    active_obj = st.session_state.get("active_obj_settings")
    if active_obj:
        from components.objective_card import _obj_actions_dialog
        _obj_actions_dialog(active_obj["id"], active_obj["title"])

    selected_team = render_sidebar()
    team_label    = selected_team if selected_team else "All"

    # Pre-compute KRs info for the AI Update dialog
    selected_quarter = st.session_state.get("selected_quarter", "Q2 2026")
    ai_objs = objectives_df
    if not objectives_df.empty:
        if "quarter" in objectives_df.columns:
            ai_objs = ai_objs[ai_objs["quarter"] == selected_quarter]
        if team_label != "All" and "sub_team" in ai_objs.columns:
            ai_objs = ai_objs[ai_objs["sub_team"] == team_label]

    if not ai_objs.empty and not krs_df.empty:
        ai_obj_ids = set(ai_objs["id"].astype(str).tolist())
        ai_krs     = krs_df[krs_df["objective_id"].astype(str).isin(ai_obj_ids)]
    else:
        ai_krs = krs_df.iloc[:0]

    # --- PRECOMPUTE KR DATA ---
    selected_week = st.session_state.get("selected_week", get_week_number())

    def _get_latest_map(df, week_limit):
        if df.empty: return {}
        rev = df[df["week_number"] <= week_limit].sort_values(["kr_id", "updated_at"], ascending=[True, False])
        return rev.drop_duplicates(subset=["kr_id"]).set_index("kr_id").to_dict("index")

    latest_map = _get_latest_map(updates_df, selected_week)
    prev_latest_map = _get_latest_map(updates_df, selected_week - 1)

    # Get all-time latest values (for Excel export) - use max week_number, not selected_week
    all_time_latest_map = _get_latest_map(updates_df, 999999) if not updates_df.empty else {}

    # ✅ FIX: Handle empty updates_df gracefully
    if not updates_df.empty and "week_number" in updates_df.columns:
        narrative_map = updates_df[updates_df["week_number"] == selected_week].sort_values(
            ["kr_id", "updated_at"], ascending=[True, False]
        ).drop_duplicates(subset=["kr_id"]).set_index("kr_id").to_dict("index")
    else:
        narrative_map = {}

    krs_info_all = {}
    krs_info_all_for_export = {}  # Separate structure with all-time latest values
    if not krs_df.empty:
        for _, kr in krs_df.iterrows():
            kid = str(kr["id"])
            latest = latest_map.get(kid)
            narrative = narrative_map.get(kid)
            prev_latest = prev_latest_map.get(kid)
            all_time_latest = all_time_latest_map.get(kid)  # Get all-time latest

            # For display (respects selected week)
            eff_val = float(latest["new_value"]) if latest else float(kr.get("current_value", 0))
            calc_kr = kr.copy(); calc_kr["current_value"] = eff_val
            pct = compute_progress(calc_kr)

            prev_pct = None
            if selected_week > 1:
                pv = float(prev_latest["new_value"]) if prev_latest else float(kr.get("current_value", 0))
                p_kr = kr.copy(); p_kr["current_value"] = pv
                prev_pct = compute_progress(p_kr)

            krs_info_all[kid] = {
                "kr": kr, "val": eff_val, "pct": pct,
                "latest": latest, "narrative": narrative, "prev_pct": prev_pct,
                "has_updates": latest is not None,
                "title": str(kr["title"]), "target": float(kr["target"]), "unit": str(kr["unit"]), # compatibility
            }

            # For Excel export (use all-time latest value)
            export_val = float(all_time_latest["new_value"]) if all_time_latest else float(kr.get("current_value", 0))
            krs_info_all_for_export[kid] = {
                "kr": kr, "val": export_val,
                "title": str(kr["title"]), "target": float(kr["target"]), "unit": str(kr["unit"]),
            }

    # Filter for Header and AI summaries based on the current team/quarter selection
    if not ai_krs.empty:
        ai_krs_ids = set(ai_krs["id"].astype(str).tolist())
        krs_info_list = [krs_info_all[kid] for kid in ai_krs_ids if kid in krs_info_all]
    else:
        krs_info_list = []
        
    st.session_state["_krs_for_ai"] = krs_info_list

    render_header(objectives_df, krs_df, updates_df, team_label, krs_info_list, krs_info_all_for_export)
    render_last_action()


    # PDF import section (sub-team views only)
    if team_label != "All":
        if st.session_state.get("show_pdf_import"):
            with st.container():
                col_title, col_exit = st.columns([3, 1])
                with col_title:
                    st.markdown("### 📄 Import OKR Update from PDF")
                with col_exit:
                    if st.button("Back to Dashboard", icon=":material/arrow_back:", width="stretch"):
                        st.session_state["show_pdf_import"] = False
                        st.session_state.pop("parsed_pdf_data", None)
                        st.rerun()
                st.caption(
                    "Upload your OKR PDF report. The system will extract the data "
                    "and show you a preview before saving anything."
                )
                pdf_file = st.file_uploader(
                    "Upload PDF", type=["pdf"], key=f"pdf_uploader_{team_label}"
                )
                if pdf_file:
                    if st.button("Parse PDF", icon=":material/search:", type="primary", key=f"parse_pdf_btn_{team_label}"):
                        with st.spinner("Reading and analyzing PDF with AI (this may take a moment)..."):
                            try:
                                parsed_data = parse_okr_pdf_with_ai(
                                    pdf_file, team_label, selected_quarter,
                                    st.secrets["OPENAI_API_KEY"],
                                )
                                st.session_state["parsed_pdf_data"] = parsed_data
                                track_action("Parsed PDF", detail=team_label)
                                st.toast("PDF successfully analyzed", icon="🔍")
                            except Exception as exc:
                                st.error(f"Error parsing PDF: {exc}")
                
                if st.session_state.get("parsed_pdf_data"):
                    render_pdf_preview_and_confirm(
                        st.session_state["parsed_pdf_data"], team_label, selected_quarter,
                    )

        # Undo button visible on main dashboard after import
        if "last_import_summary" in st.session_state:
            import_summary = st.session_state["last_import_summary"]
            import_time = import_summary.get("timestamp", 0)
            
            # Only show the button if it's less than 15 seconds old
            if time.time() - import_time < 60:
                st.warning("⏪ Recent import detected. You can revert it if needed.")
                c1, c2 = st.columns([1, 4])
                with c1:
                    if st.button("Undo", icon=":material/undo:", width="stretch", help="Revert the last PDF import changes"):
                        undo_last_import(import_summary)
                        st.session_state.pop("last_import_summary", None)
                        st.rerun()
                with c2:
                    if st.button("Dismiss", icon=":material/check:", width="stretch", help="Keep changes and hide this message"):
                        st.session_state.pop("last_import_summary", None)
                        st.rerun()
            else:
                st.session_state.pop("last_import_summary", None)

    # Filter display objectives by quarter + team
    display_objs = objectives_df
    if not objectives_df.empty and "quarter" in objectives_df.columns:
        display_objs = display_objs[display_objs["quarter"] == selected_quarter]

    # ✅ DYNAMIC ORDERING: Sort by subteam if viewing "All"
    if team_label == "All" and not display_objs.empty and "sub_team" in display_objs.columns:
        # Define the subteam order
        subteam_order = {
            "New Initiatives": 0,
            "FinOps": 1,
            "AI Experience": 2,
            "AI Monetization": 3,
            "Security & Compliance Ops": 4,
        }
        # Map subteam to order, use 999 for unknown teams (put them at the end)
        display_objs = display_objs.copy()
        display_objs["_order"] = display_objs["sub_team"].map(lambda x: subteam_order.get(x, 999))
        # Sort by order, preserving the original row order within each subteam
        display_objs = display_objs.sort_values("_order", kind="stable")
        display_objs = display_objs.drop(columns=["_order"])
    elif team_label != "All" and not display_objs.empty and "sub_team" in display_objs.columns:
        display_objs = display_objs[display_objs["sub_team"] == team_label]

    if display_objs.empty:
        st.info(
            f"No OKRs found for **{selected_quarter}** · **{team_label}**. "
            "Please create one or import from PDF."
        )
    else:
        # Presentation mode toggle (only for subteams, not for "All") - now in header
        presentation_key = f"presentation_mode_{team_label}"
        is_presentation_mode = st.session_state.get(presentation_key, False)

        # Render objectives

        for i, (_, obj_row) in enumerate(display_objs.iterrows()):
            render_objective_card(
                obj_row,
                krs_df,
                updates_df,
                krs_info_all,
                is_primary=(i == 0),
                read_only=is_presentation_mode  # Pass read_only flag based on presentation mode
            )


    if team_label != "All":
        presentation_key = f"presentation_mode_{team_label}"
        is_presentation_mode = st.session_state.get(presentation_key, False)

        # Only show "Add Objective" button in edit mode
        if not is_presentation_mode:
            if st.button("Add Objective", icon=":material/add:", key=f"add_obj_{team_label}", type="secondary"):
                add_objective_dialog(team_label, selected_quarter)

    # Activity log for "All" view (sub-team views have it at the bottom)
    if team_label == "All":
        st.divider()
        render_activity_log()

    # Weekly Notes + Charts section (sub-team views only)
    if team_label != "All":
        week_number   = st.session_state.get("selected_week", get_week_number())
        is_past_week = week_number < get_week_number()

        existing_note = get_weekly_note(team_label, selected_quarter, week_number)
        note_content  = existing_note.get("content", "")
        current_charts = get_weekly_charts(team_label, selected_quarter, week_number)

        # Only show sections if they have content (in presentation mode)
        is_presentation = st.session_state.get(f"presentation_mode_{team_label}", False)

        # Show Weekly Notes section only if there's content (or not in presentation mode)
        if note_content or not is_presentation:
            # Section Header
            st.markdown('<div style="font-size:18px; font-weight:700; color:#fff; border-bottom:1px solid rgba(255,255,255,0.06); padding-bottom:8px; margin-bottom:4px; margin-top:24px;">📝 Additional Weekly Notes</div>', unsafe_allow_html=True)
            st.markdown('<div style="font-size:13px; color:rgba(255,255,255,0.4); margin-bottom:16px;">Important context and decisions for this period</div>', unsafe_allow_html=True)

            # Display Mode (Card)
            if note_content:
                formatted_note = note_content.replace('\n', '<br>')
                st.markdown(f"""
                <div style="background-color: rgba(255,255,255,0.03); padding: 1.2rem; border-radius: 12px; border: 1px solid rgba(255,255,255,0.08); margin-bottom: 1rem; color: rgba(255,255,255,0.85); font-size: 0.95rem; line-height:1.6;">
                    {formatted_note}
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"""
                <div style="padding: 24px; text-align:center; background:rgba(255,255,255,0.02); border-radius:12px; border:1px dashed rgba(255,255,255,0.06); color:rgba(255,255,255,0.3); font-size:14px; margin-bottom:16px;">
                    No narrative notes recorded for Week {week_number}.
                </div>
                """, unsafe_allow_html=True)

            # Edit Section (hide in presentation mode)
            if not is_presentation:
                with st.expander("✏️ Edit Weekly Note"):
                    note_text = st.text_area(
                        "Note Content",
                        value=note_content,
                        height=200,
                        key=f"editor_{team_label}_{week_number}",
                        label_visibility="collapsed"
                    )
                    col1, col2 = st.columns([4, 1])
                    with col2:
                        if st.button("Save Changes", width="stretch", type="primary", key=f"btn_save_{team_label}_{week_number}"):
                            email = st.session_state.get("user", {}).get("email", "unknown")
                            with st.spinner("Saving..."):
                                save_weekly_note(team_label, selected_quarter, week_number, note_text, email)
                                st.cache_data.clear()
                                track_action("Saved weekly note", detail=f"{team_label} W{week_number}")
                                st.toast("Changes saved!", icon="✅")
                                st.rerun()

        # Charts Section - only show if there are charts (or not in presentation mode)
        if current_charts or not is_presentation:
            st.markdown('<div style="font-size:18px; font-weight:700; color:#fff; border-bottom:1px solid rgba(255,255,255,0.06); padding-bottom:8px; margin-bottom:4px; margin-top:32px;">📊 Charts & Screenshots</div>', unsafe_allow_html=True)
            st.markdown('<div style="font-size:13px; color:rgba(255,255,255,0.4); margin-bottom:16px;">Evidence and visual data for the weekly sync</div>', unsafe_allow_html=True)

            # File uploader (hide in presentation mode)
            if not is_presentation:
                uploaded_files = st.file_uploader(
                    "Upload charts",
                    type=["png", "jpg", "jpeg"],
                    accept_multiple_files=True,
                    label_visibility="collapsed",
                    key=f"chart_uploader_{team_label}_{week_number}",
                )
                if uploaded_files:
                    if st.button("Upload to Drive", type="primary", key=f"upload_charts_{team_label}_{week_number}"):
                        with st.spinner("Uploading images to Google Drive..."):
                            email = st.session_state.get("user", {}).get("email", "unknown")
                            upload_charts_to_drive(uploaded_files, team_label, selected_quarter, week_number, email)
                            track_action("Uploaded charts", detail=f"{len(uploaded_files)} file(s)")
                            st.toast("✅ Images uploaded successfully", icon="📊")
                        st.rerun()

            if current_charts:
                st.markdown("**This week's charts:**")
                chart_cols = st.columns(min(len(current_charts), 2))
                for i, chart in enumerate(current_charts):
                    with chart_cols[i % 2]:
                        try:
                            img_bytes = download_drive_file(chart["drive_file_id"])
                            st.image(img_bytes, caption=chart["filename"], width="stretch")
                        except Exception as e:
                            st.error(f"Error loading image: {e}")

                        # Delete button (hide in presentation mode)
                        if not is_presentation:
                            if st.button("Remove", icon=":material/delete:", key=f"del_chart_{chart['id']}", type="secondary"):
                                with st.spinner("Deleting image..."):
                                    delete_chart_from_drive(str(chart["id"]))
                                    track_action("Deleted chart", detail=chart.get("filename", ""))
                                    st.toast("Image deleted", icon="🗑️")
                                st.rerun()

        # Activity log at the bottom of sub-team views
        st.divider()
        render_activity_log()


# ---------------------------------------------------------------------------
if DEV_MODE:
    if st.session_state.get("dev_authenticated"):
        render_dashboard()
    else:
        render_login_page()
else:
    if require_login():
        render_dashboard()
    else:
        render_login_page()
